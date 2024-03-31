from django.core.management.base import BaseCommand
from feed.models import Tempaper

import os
import openpyxl

from utils import database, debug, common

BRAND = "Tempaper"
FILEDIR = f"{os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}/files"


class Command(BaseCommand):
    help = f"Build {BRAND} Database"

    def add_arguments(self, parser):
        parser.add_argument('functions', nargs='+', type=str)

    def handle(self, *args, **options):
        if "feed" in options['functions']:
            common.downloadFileFromSFTP(
                brand=BRAND,
                src="/tempaper/datasheets/tempaper-master.xlsx",
                dst=f"{FILEDIR}/tempaper-master.xlsx",
                fileSrc=True,
                delete=False
            )

            processor = Processor()
            feeds = processor.fetchFeed()
            processor.DatabaseManager.writeFeed(feeds=feeds)

        if "validate" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.validateFeed()

        if "status" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.statusSync(fullSync=False)

        if "content" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.contentSync(private=True)

        if "price" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.priceSync()

        if "tag" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.tagSync()

        if "add" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.addProducts(private=True)

        if "update" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.updateProducts(
                feeds=Tempaper.objects.all(), private=True)

        if "image" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.downloadImages()

        if "inventory" in options['functions']:
            common.downloadFileFromSFTP(
                brand=BRAND,
                src="/tempaper/datasheets/tempaper-master.xlsx",
                dst=f"{FILEDIR}/tempaper-master.xlsx",
                fileSrc=True,
                delete=False
            )

            processor = Processor()
            stocks = processor.inventory()
            processor.DatabaseManager.updateInventory(
                stocks=stocks, type=1, reset=True)


class Processor:
    def __init__(self):
        self.DatabaseManager = database.DatabaseManager(
            brand=BRAND, Feed=Tempaper)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def fetchFeed(self):
        # Get Product Feed
        products = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/tempaper-master.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            try:
                # Primary Keys
                mpn = common.toText(row[3])
                sku = f"TP {mpn}"

                pattern = common.toText(row[4])
                color = common.toText(row[5])

                name = common.toText(row[8])

                # Categorization
                brand = BRAND
                type = common.toText(row[0])
                manufacturer = brand
                collection = common.toText(row[2].replace("Tempaper", ""))

                # Main Information
                description = common.toText(row[9])

                width = common.toFloat(row[17])
                length = common.toFloat(row[18]) * 12

                # Additional Information
                yardsPR = common.toInt(row[14])
                match = common.toText(row[25])
                material = common.toText(row[27])
                care = common.toText(row[32])
                weight = common.toFloat(row[22])
                country = common.toText(row[33])

                coverage = common.toText(row[21])
                specs = [
                    ("Coverage", coverage),
                ]

                features = []
                for id in range(28, 30):
                    if row[id]:
                        features.append(common.toText(row[id]))

                # Measurement
                uom = f"{common.toText(row[13])}"

                # Pricing
                cost = common.toFloat(row[10])
                map = common.toFloat(row[11])

                # Tagging
                keywords = f"{collection} {pattern} {description} {material} {match} {row[28]} {row[29]}"
                colors = color

                # Image
                thumbnail = common.toText(row[34]).replace("dl=0", "dl=1")

                roomsets = []
                for id in range(35, 39):
                    if row[id]:
                        roomsets.append(common.toText(
                            row[id]).replace("dl=0", "dl=1"))

                # Status
                statusP = True
                if type == "Wallpaper":
                    statusS = True
                else:
                    statusS = False

                # Fine-tuning
                name = f"{collection} {pattern} {color} {type}"

                # Exceptions
                if cost == 0 or not pattern or not color or not type:
                    continue

            except Exception as e:
                debug.warn(BRAND, str(e))
                continue

            product = {
                'mpn': mpn,
                'sku': sku,
                'pattern': pattern,
                'color': color,
                'name': name,

                'brand': brand,
                'type': type,
                'manufacturer': manufacturer,
                'collection': collection,

                'description': description,
                'width': width,
                'length': length,

                'yardsPR': yardsPR,
                'material': material,
                'match': match,
                'care': care,
                'country': country,
                'weight': weight,

                'specs': specs,
                'features': features,

                'uom': uom,

                'cost': cost,
                'map': map,

                'keywords': keywords,
                'colors': colors,

                'thumbnail': thumbnail,
                'roomsets': roomsets,

                'statusP': statusP,
                'statusS': statusS,
            }
            products.append(product)

        return products

    def inventory(self):
        stocks = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/tempaper-master.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = common.toText(row[3])

            try:
                product = Tempaper.objects.get(mpn=mpn)
            except Tempaper.DoesNotExist:
                continue

            sku = product.sku
            stockP = common.toInt(row[6])

            stock = {
                'sku': sku,
                'quantity': stockP,
                'note': ""
            }
            stocks.append(stock)

        return stocks
