from django.core.management.base import BaseCommand
from feed.models import Brewster

import os
import openpyxl
import paramiko
import csv
import codecs

from utils import database, debug, common, const
from vendor.models import Product

BRAND = "Brewster"
FILEDIR = f"{os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}/files"
IMAGEDIR = f"{os.path.expanduser('~')}/admin/vendor/management/files/images"


class Command(BaseCommand):
    help = f"Build {BRAND} Database"

    def add_arguments(self, parser):
        parser.add_argument('functions', nargs='+', type=str)

    def handle(self, *args, **options):
        if "feed" in options['functions']:
            processor = Processor()
            processor.downloadDatasheets()
            feeds = processor.fetchFeed()
            processor.DatabaseManager.writeFeed(feeds=feeds)

        if "validate" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.validateFeed()

        if "status" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.statusSync(fullSync=False)

        if "content" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.contentSync()

        if "price" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.priceSync()

        if "tag" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.tagSync()

        if "add" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.addProducts()

        if "image" in options['functions']:
            processor = Processor()
            processor.image()

        if "inventory" in options['functions']:
            common.downloadFileFromSFTP(
                brand=BRAND,
                src="/brewster",
                dst=f"{FILEDIR}/brewster-inventory.csv",
                fileSrc=False,
                delete=True
            )

            processor = Processor()
            stocks = processor.inventory()
            processor.DatabaseManager.updateInventory(
                stocks=stocks, type=1, reset=True)


class Processor:
    def __init__(self):
        self.DatabaseManager = database.DatabaseManager(
            brand=BRAND, Feed=Brewster)

        transport = paramiko.Transport(
            (const.sftp[f"{BRAND} Images"]["host"], const.sftp[f"{BRAND} Images"]["port"]))
        transport.connect(
            username=const.sftp[f"{BRAND} Images"]["user"], password=const.sftp[f"{BRAND} Images"]["pass"])
        self.imageServer = paramiko.SFTPClient.from_transport(transport)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def downloadDatasheets(self):
        self.imageServer.chdir(path='/WallpaperBooks')
        collections = self.imageServer.listdir()

        for collection in collections:
            if "All Wallpaper Images" in collection:
                continue

            collectionDir = self.imageServer.listdir(collection)
            exclusions = {"$", "~", "TheCottageData",
                          "Advantage Neutral Black White"}
            xlsx_files = [file for file in collectionDir if file.endswith(
                '.xlsx') and not any(excl in file for excl in exclusions)]

            for filename in xlsx_files:
                try:
                    remote_path = f"{collection}/{filename}"
                    local_path = os.path.join(
                        FILEDIR, "brewster", f"{collection}.xlsx")
                    self.imageServer.get(remote_path, local_path)
                    debug.log(
                        BRAND, f"Downloaded {filename} from Brewster SFTP")
                except Exception as e:
                    debug.warn(
                        BRAND, f"Downloading {filename} from Brewster SFTP has been failed. Error: {e}")
                    continue

            if not xlsx_files or 'TheCottageData.xlsx' in xlsx_files and len(xlsx_files) == 1:
                debug.warn(
                    BRAND, f"No datasheets found in {collection} directory")

    def fetchFeed(self):
        # Price & Discontinued
        prices = {}
        discontinuedMPNs = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/brewster-price.xlsx", data_only=True)
        sh = wb.worksheets[0]
        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = common.toText(row[0])

            if row[14] == "Y":
                discontinuedMPNs.append(mpn)

            cost = common.toFloat(row[13])
            map = common.toFloat(row[12])
            msrp = common.toFloat(row[11])

            prices[mpn] = {
                'cost': cost,
                'map': map,
                'msrp': msrp
            }

        # Best Sellers
        bestsellingMPNs = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/brewster-bestsellers.xlsx", data_only=True)
        sh = wb.worksheets[0]
        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = common.toText(row[0])
            bestsellingMPNs.append(mpn)

        # Get Product Feed
        products = []

        def readDatasheet(sh):
            data = [row for row in sh.iter_rows(values_only=True)]

            headers = data[0] if data else []

            data_headers = []
            for header in headers:
                if header == 'Pattern' and 'MPN' not in data_headers:
                    data_headers.append('MPN')
                else:
                    data_headers.append(header)

            data_rows = data[1:] if len(data) > 1 else []

            return (data_headers, data_rows)

        for datasheet in os.listdir(f"{FILEDIR}/brewster"):
            wb = openpyxl.load_workbook(
                f"{FILEDIR}/brewster/{datasheet}", data_only=True)
            sh = wb.worksheets[0]

            headers, data = readDatasheet(sh=sh)

            header_to_id_map = {
                "Book Name": "collectionId",
                "Brand": "manufacturerId",
                "MPN": "mpnId",
                "Name": "nameId",
                "Product Type": "usageId",
                "Description": "descriptionId",
                "MSRP": "msrpId",
                "MAP": "mapId",
                "Barcode": "upcId",
                "Width (in)": "widthId",
                "Length (ft)": "lengthId",
                "Coverage": "coverageId",
                "Unit Weight": "weightId",
                "Repeat (in)": "repeatId",
                "Match": "matchId",
                "Paste": "pasteId",
                "Material": "materialId",
                "Washability": "washId",
                "Removability": "removeId",
                "Colorway": "colorId",
                "Color Family": "colorsId",
                "Style": "styleId",
                "Pattern": "patternId",
                "Theme": "themeId",
                "Country of Origin": "countryId",
            }

            ids = {id_key: -1 for id_key in header_to_id_map.values()}

            for colId, header in enumerate(headers):
                if header in header_to_id_map:
                    id_key = header_to_id_map[header]
                    ids[id_key] = colId

            for dataRow in data:
                row = dataRow + (None,)
                try:
                    # Primary Attributes
                    mpn = common.toText(row[ids["mpnId"]])

                    sku_prefix = "STREET" if row[ids["manufacturerId"]
                                                 ] == "A-Street Prints" else "BREWSTER"
                    sku = f"{sku_prefix} {mpn}"

                    pattern = common.toText(row[ids["patternId"]]) or mpn
                    color = common.toText(row[ids["colorId"]])

                    # Categorization
                    brand = BRAND
                    type = "Wallpaper"

                    manufacturer = row[ids["manufacturerId"]]

                    collection = f'{manufacturer} {row[ids["collectionId"]]}'

                    # Main Information
                    description = common.toText(row[ids["descriptionId"]])
                    width = common.toFloat(row[ids["widthId"]])
                    length = common.toFloat(row[ids["lengthId"]]) * 12
                    repeat = common.toFloat(row[ids["repeatId"]])

                    # Additional Information
                    yardsPR = common.toFloat(length / 36)
                    match = common.toText(row[ids["matchId"]])
                    material = common.toText(row[ids["materialId"]])
                    usage = common.toText(row[ids["usageId"]])
                    weight = common.toFloat(row[ids["weightId"]]) or 1
                    country = common.toText(row[ids["countryId"]])
                    upc = common.toText(row[ids["upcId"]])

                    coverage = common.toText(row[ids["coverageId"]])
                    paste = common.toText(row[ids["pasteId"]])
                    wash = common.toText(row[ids["washId"]])
                    remove = common.toText(row[ids["removeId"]])
                    specs = [
                        ("Coverage", coverage),
                        ("Paste", paste),
                        ("Washability", wash),
                        ("Removability", remove),
                    ]

                    # Measurement
                    uom = "Roll"

                    # Pricing
                    msrp = common.toFloat(
                        row[ids["msrpId"]]) if ids["msrpId"] > 0 else 0
                    cost = msrp / 2

                    map = common.toFloat(
                        row[ids["mapId"]]) if ids["mapId"] > 0 else 0

                    if mpn in prices:
                        cost = prices[mpn]['cost']
                        map = prices[mpn]['map']
                        msrp = prices[mpn]['msrp']

                    # Tagging
                    keywords = f'{collection} {pattern} {description} {row[ids["styleId"]]} {row[ids["themeId"]]}'
                    colors = common.toText(row[ids["colorsId"]])

                    # Status
                    statusP = collection not in [
                        'Scalamandre', 'Eijffinger', 'Eiffinger'] and mpn not in discontinuedMPNs
                    statusS = False

                    bestSeller = mpn in bestsellingMPNs

                    # Fine-tuning
                    if not color:
                        color = colors
                    else:
                        colors = f"{color} {colors}"

                    custom = {
                        'originalBrand': manufacturer
                    }
                    manufacturer = "Brewster Home Fashions" if manufacturer != "A-Street Prints" else "A-Street Prints"
                    name = common.toText(
                        row[ids["nameId"]]) or f"{pattern} {color} {type}"

                    if "Mural" in usage:
                        type = "Mural"
                        uom = "Each"

                    pattern = name.replace(color, "").replace("Wallpaper", "").replace(
                        "Mural", "").replace("Border", "").replace("  ", " ").strip()

                    # Exceptions
                    if cost == 0 or not pattern or not color or not type:
                        continue

                except Exception as e:
                    debug.warn(BRAND, str(e))
                    continue

                product = {
                    'mpn': mpn,
                    'sku': sku,
                    'pattern': pattern,
                    'color': color,
                    'name': name,

                    'brand': brand,
                    'type': type,
                    'manufacturer': manufacturer,
                    'collection': collection,

                    'description': description,
                    'width': width,
                    'length': length,
                    'repeatV': repeat,

                    'yardsPR': yardsPR,
                    'match': match,
                    'material': material,
                    'weight': weight,
                    'usage': usage,
                    'country': country,
                    'upc': upc,

                    'specs': specs,

                    'uom': uom,

                    'cost': cost,
                    'map': map,

                    'keywords': keywords,
                    'colors': colors,

                    'statusP': statusP,
                    'statusS': statusS,
                    'bestSeller': bestSeller,

                    'custom': custom,
                }
                products.append(product)

        return products

    def image(self, fullSync=False):
        has_image_ids = set(Product.objects.filter(manufacturer__brand=BRAND)
                                           .filter(images__position=1)
                                           .values_list('shopifyId', flat=True)
                                           .distinct())

        feeds = Brewster.objects.exclude(productId=None)
        if not fullSync:
            feeds = feeds.exclude(productId__in=has_image_ids)

        for feed in feeds.iterator():
            collection = feed.collection.strip()
            mpn = feed.mpn
            product_id = feed.productId
            original_brand = feed.custom.get('originalBrand')

            if original_brand != "Advantage" and collection != "Eijffinger Web Only":
                collection = collection.replace(original_brand, "").strip()

            for image_size in ['Images/300dpi', 'Images/72dpi', 'Images', '300dpi', '72dpi']:
                try:
                    dir_path = f'/WallpaperBooks/{collection}'
                    if image_size:
                        dir_path += f'/{image_size}'
                    self.imageServer.chdir(path=dir_path)
                    break
                except Exception as e:
                    pass

            files = set(self.imageServer.listdir())

            for index, suffix in enumerate(["", "_Room", "_Room_2", "_Room_3", "_Room_4"], start=1):
                file_name = f"{mpn}{suffix}.jpg"
                if file_name in files:
                    local_path = f"{IMAGEDIR}/thumbnail/{product_id}.jpg" if index == 1 else f"{IMAGEDIR}/roomset/{product_id}_{index}.jpg"
                    self.imageServer.get(file_name, local_path)
                    debug.log(BRAND, f"downloaded {local_path}")

    def inventory(self):
        stocks = []

        f = open(f"{FILEDIR}/brewster-inventory.csv", "rb")
        cr = csv.reader(codecs.iterdecode(f, encoding="ISO-8859-1"))
        for row in cr:
            mpn = row[0]

            try:
                product = Brewster.objects.get(mpn=mpn)
            except Brewster.DoesNotExist:
                continue

            sku = product.sku
            stockP = common.toInt(row[3])

            stock = {
                'sku': sku,
                'quantity': stockP,
                'note': ""
            }
            stocks.append(stock)

        return stocks
