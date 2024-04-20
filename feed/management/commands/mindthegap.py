from django.core.management.base import BaseCommand
from feed.models import MindTheGap

import os
import environ
import openpyxl
import re

from utils import database, debug, common

env = environ.Env()

BRAND = "MindTheGap"
FILEDIR = f"{os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}/files"


class Command(BaseCommand):
    help = f"Build {BRAND} Database"

    def add_arguments(self, parser):
        parser.add_argument('functions', nargs='+', type=str)

    def handle(self, *args, **options):
        if "feed" in options['functions']:
            processor = Processor()
            feeds = processor.fetchFeed()
            processor.DatabaseManager.writeFeed(feeds=feeds)

        if "validate" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.validateFeed()

        if "status" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.statusSync(fullSync=False)

        if "content" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.contentSync()

        if "price" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.priceSync()

        if "tag" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.tagSync()

        if "add" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.addProducts()

        if "image" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.downloadImages()

        if "inventory" in options['functions']:
            common.downloadFileFromSFTP(
                brand=BRAND,
                src="/mindthegap/Inventory/MINDTHEGAP STOCK cushions.xlsx",
                dst=f"{FILEDIR}/mindthegap-pillow-inventory.xlsx",
                fileSrc=True,
                delete=False
            )

            common.downloadFileFromSFTP(
                brand=BRAND,
                src="/mindthegap/Inventory/MINDTHEGAP STOCK fabrics.xlsx",
                dst=f"{FILEDIR}/mindthegap-fabric-inventory.xlsx",
                fileSrc=True,
                delete=False
            )

            processor = Processor()
            stocks = processor.inventory()
            processor.DatabaseManager.updateInventory(
                stocks=stocks, type=1, reset=True)


class Processor:
    def __init__(self):
        self.DatabaseManager = database.DatabaseManager(
            brand=BRAND, Feed=MindTheGap)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def fetchFeed(self):
        # Get Product Feed
        products = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/mindthegap-master.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            try:
                # Primary Keys
                mpn = common.toText(row[2])
                sku = f"MTG {mpn}"

                pattern = common.toText(row[3])
                color = common.toText(row[4])

                # Categorization
                brand = BRAND
                manufacturer = BRAND
                type = common.toText(row[0]).title()
                collection = common.toText(row[1])

                # Main Information
                description = common.toText(
                    f"{row[12]} {row[11]}" if row[11] else row[12])

                size = common.toText(row[7]).replace(',', '.')

                if row[18] and '/' in row[18]:
                    repeat = row[18].split('/')[1]
                else:
                    repeat = ""

                specs = [("Packing", "3 rolls in a box")
                         ] if "3*" in size else []

                # Additional Information
                usage = common.toText(row[0] or row[17])
                material = common.toText(row[14])
                finish = common.toText(row[15])
                care = common.toText(row[20])
                country = common.toText(row[19])
                weight = common.toFloat(row[22]) * 2.2
                upc = common.toInt(row[23])

                # Measurement
                uom = type

                # Pricing
                cost = common.toFloat(row[10])

                # Tagging
                keywords = f"{collection} {pattern} {description} {color} {row[13]} {material} {finish}"
                colors = row[5]

                # Status
                statusP = type != "Fabric"
                statusS = False

                # Fine-tuning
                TYPE_DICT = {
                    "Fabrics": "Fabric",
                    "Designer Wallpaper": "Wallpaper",
                    "Metallic Wallpaper": "Wallpaper",
                    "Complementary Wallpaper": "Wallpaper",
                }
                type = TYPE_DICT.get(type, type)

                UOM_DICT = {
                    "Wallpaper": "Roll",
                    "Fabric": "Yard",
                    "Pillow": "Item"
                }
                uom = UOM_DICT.get(type, uom)
                uom = "3 rolls in a box" if "3*" in size else uom

                sizePattern = re.findall(
                    r"(\d+(?:\.\d+)?)x(\d+(?:\.\d+)?)", size)
                width = common.toFloat(sizePattern[0][0]) if sizePattern else 0
                length = common.toFloat(
                    sizePattern[0][1]) if sizePattern else 0

                yardsPR = common.toFloat(length / 36)

                size = f'{width}" x {length}"'

                repeatPattern = re.findall(
                    r"W(\d+(?:\.\d+)?)in\s*x\s*L(\d+(?:\.\d+)?)in|(\d+(?:\.\d+)?)in", repeat)
                repeatH = common.toFloat(
                    repeatPattern[0][0]) if repeatPattern else 0
                repeatV = common.toFloat(
                    repeatPattern[0][1]) if repeatPattern else 0

                name = f"{pattern} {color} {type}"

                # Exceptions
                if cost == 0 or not pattern or not color or not type:
                    continue

            except Exception as e:
                debug.warn(BRAND, str(e))
                continue

            product = {
                'mpn': mpn,
                'sku': sku,
                'pattern': pattern,
                'color': color,
                'name': name,

                'brand': brand,
                'type': type,
                'manufacturer': manufacturer,
                'collection': collection,

                'description': description,
                'width': width,
                'length': length,
                'size': size,
                'repeatV': repeatV,
                'repeatH': repeatH,

                'usage': usage,
                'yardsPR': yardsPR,
                'material': material,
                'finish': finish,
                'care': care,
                'country': country,
                'weight': weight,
                'upc': upc,

                'specs': specs,

                'uom': uom,

                'cost': cost,

                'keywords': keywords,
                'colors': colors,

                'statusP': statusP,
                'statusS': statusS,
            }
            products.append(product)

        return products

    def inventory(self):
        stocks = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/mindthegap-pillow-inventory.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = common.toText(row[1])

            try:
                product = MindTheGap.objects.get(mpn=mpn)
            except MindTheGap.DoesNotExist:
                continue

            sku = product.sku
            stockP = common.toInt(row[5])

            stock = {
                'sku': sku,
                'quantity': stockP,
                'note': ""
            }
            stocks.append(stock)

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/mindthegap-fabric-inventory.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = common.toText(row[1])

            try:
                product = MindTheGap.objects.get(mpn=mpn)
            except MindTheGap.DoesNotExist:
                continue

            sku = product.sku
            stockP = common.toInt(row[7])

            stock = {
                'sku': sku,
                'quantity': stockP,
                'note': ""
            }
            stocks.append(stock)

        wallpapers = MindTheGap.objects.filter(type="Wallpaper")
        for wallpaper in wallpapers:
            stock = {
                'sku': wallpaper.sku,
                'quantity': 5,
                'note': ""
            }
            stocks.append(stock)

        return stocks
