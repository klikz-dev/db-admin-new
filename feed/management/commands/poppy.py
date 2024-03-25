from django.core.management.base import BaseCommand
from feed.models import Poppy

import os
import environ
import openpyxl

from utils import database, debug, common

env = environ.Env()

BRAND = "Poppy"
FILEDIR = f"{os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}/files"


class Command(BaseCommand):
    help = f"Build {BRAND} Database"

    def add_arguments(self, parser):
        parser.add_argument('functions', nargs='+', type=str)

    def handle(self, *args, **options):
        if "feed" in options['functions']:
            processor = Processor()
            feeds = processor.fetchFeed()
            processor.DatabaseManager.writeFeed(feeds=feeds)

        if "validate" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.validateFeed()

        if "status" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.statusSync(fullSync=True)

        if "content" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.contentSync()

        if "price" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.priceSync()

        if "tag" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.tagSync()

        if "add" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.addProducts()

        if "update" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.updateProducts(
                feeds=Poppy.objects.all())

        if "image" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.downloadImages()

        if "inventory" in options['functions']:
            common.downloadFileFromSFTP(
                brand=BRAND,
                src="/poppy/Decorators Best - Poppy Inventory.xlsx",
                dst=f"{FILEDIR}/poppy-inventory.xlsx",
                fileSrc=True,
                delete=False
            )

            processor = Processor()
            stocks = processor.inventory()
            processor.DatabaseManager.updateInventory(
                stocks=stocks, type=1, reset=True)


class Processor:
    def __init__(self):
        self.DatabaseManager = database.DatabaseManager(
            brand=BRAND, Feed=Poppy)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def fetchFeed(self):
        # Get Product Feed
        products = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/poppy-master.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=3, values_only=True):
            try:
                # Primary Keys
                mpn = common.toText(row[2])
                sku = f"PP {mpn}"

                pattern = common.toText(row[3])
                color = common.toText(row[4])

                # Categorization
                brand = BRAND
                manufacturer = "Poppy Print Studio"

                type = "Wallpaper"

                collection = common.toText(row[1])

                # Main Information
                description = common.toText(row[8])

                width = common.toFloat(row[16])
                length = common.toFloat(row[17]) * 12

                repeatV = common.toFloat(row[20])
                repeatH = common.toFloat(row[21])

                # Additional Information
                yardsPR = common.toInt(row[13])
                match = common.toText(row[22])
                material = common.toText(row[24])
                weight = common.toFloat(row[19])
                country = common.toText(row[29])

                coverage = common.toText(row[18])
                paste = common.toText(row[23])
                washability = common.toText(row[25])
                removability = common.toText(row[26])
                specs = [
                    ("Coverage", coverage),
                    ("Paste", paste),
                    ("Washability", washability),
                    ("Removability", removability),
                ]

                # Measurement
                uom = common.toText(row[12])

                # Pricing
                cost = common.toFloat(row[9])
                map = common.toFloat(row[10])

                # Tagging
                keywords = f"{match} {paste} {material} {washability} {removability} {common.toText(row[27])} {collection} {pattern} {description}"
                colors = color

                # Image
                if row[30]:
                    thumbnail = f'https://drive.google.com/u/0/uc?id={row[30].replace("https://drive.google.com/file/d/", "").split("/")[0]}&export=download'
                else:
                    thumbnail = ""

                roomsets = []
                for id in range(31, 35):
                    if row[id]:
                        roomset = f'https://drive.google.com/u/0/uc?id={row[id].replace("https://drive.google.com/file/d/", "").split("/")[0]}&export=download'
                        roomsets.append(roomset)

                # Status
                statusP = True
                statusS = True

                # Fine-tuning
                name = f"{pattern} {color} {type}"

                # Exceptions
                if cost == 0 or not pattern or not color or not type:
                    continue

            except Exception as e:
                debug.warn(BRAND, str(e))
                continue

            product = {
                'mpn': mpn,
                'sku': sku,
                'pattern': pattern,
                'color': color,
                'name': name,

                'brand': brand,
                'type': type,
                'manufacturer': manufacturer,
                'collection': collection,

                'description': description,
                'width': width,
                'length': length,
                'repeatV': repeatV,
                'repeatH': repeatH,

                'yardsPR': yardsPR,
                'match': match,
                'material': material,
                'weight': weight,
                'country': country,

                'specs': specs,

                'uom': uom,

                'cost': cost,
                'map': map,

                'keywords': keywords,
                'colors': colors,

                'thumbnail': thumbnail,
                'roomsets': roomsets,

                'statusP': statusP,
                'statusS': statusS,
            }
            products.append(product)

        return products

    def inventory(self):
        stocks = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/poppy-inventory.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = common.toText(row[2])

            try:
                product = Poppy.objects.get(mpn=mpn)
            except Poppy.DoesNotExist:
                continue

            sku = product.sku
            stockP = common.toInt(row[5])

            stock = {
                'sku': sku,
                'quantity': stockP,
                'note': ""
            }
            stocks.append(stock)

        return stocks
