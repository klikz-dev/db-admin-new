from django.core.management.base import BaseCommand
from feed.models import PeninsulaHome

import os
import openpyxl

from utils import database, debug, common

BRAND = "Peninsula Home"
FILEDIR = f"{os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}/files"


class Command(BaseCommand):
    help = f"Build {BRAND} Database"

    def add_arguments(self, parser):
        parser.add_argument('functions', nargs='+', type=str)

    def handle(self, *args, **options):
        if "feed" in options['functions']:
            processor = Processor()
            feeds = processor.fetchFeed()
            processor.DatabaseManager.writeFeed(feeds=feeds)

        if "validate" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.validateFeed()

        if "status" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.statusSync(fullSync=False)

        if "content" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.contentSync()

        if "price" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.priceSync()

        if "tag" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.tagSync()

        if "add" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.addProducts()

        if "image" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.downloadImages()

        if "inventory" in options['functions']:
            files = common.browseSFTP(brand=BRAND, src="/peninsulahome/")
            for file in files:
                if "xlsx" in file:
                    common.downloadFileFromSFTP(
                        brand=BRAND,
                        src=f"/peninsulahome/{file}",
                        dst=f"{FILEDIR}/peninsulahome-inventory.xlsx",
                        fileSrc=True,
                        delete=True
                    )

            processor = Processor()
            stocks = processor.inventory()
            processor.DatabaseManager.updateInventory(
                stocks=stocks, type=1, reset=True)


class Processor:
    def __init__(self):
        self.DatabaseManager = database.DatabaseManager(
            brand=BRAND, Feed=PeninsulaHome)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def fetchFeed(self):
        # Stocks & discontinued
        stocks = {}
        discontinuedMPNs = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/peninsulahome-inventory.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = common.toText(row[0])

            stockP = common.toFloat(row[1])
            stockNote = common.toText(row[2])

            stocks[mpn] = (stockP, stockNote)

        sh = wb.worksheets[1]

        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = common.toText(row[0])
            discontinuedMPNs.append(mpn)

        # Get Product Feed
        products = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/peninsulahome-master.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=3, values_only=True):
            try:
                # Primary Keys
                mpn = common.toText(row[0])
                sku = f"PH {mpn}"

                pattern = common.toText(row[1]).replace(" ,", ",")
                color = common.toText(row[13])

                name = pattern

                # Categorization
                brand = BRAND
                type = "Furniture"
                manufacturer = brand
                collection = common.toText(row[2])

                # Main Information
                description = common.toText(row[12])
                width = common.toFloat(row[9])
                length = common.toFloat(row[8])
                height = common.toFloat(row[10])

                # Additional Information
                material = common.toText(row[13])
                country = common.toText(row[15])
                weight = common.toFloat(row[7])

                dimension = common.toText(row[11])
                specs = [
                    ("Dimension", dimension),
                ]

                features = [
                    f"Fabric: {common.toText(row[14])}"] if row[14] else []

                # Measurement
                uom = f"Item"

                # Pricing
                cost = common.toFloat(row[3])
                map = common.toFloat(row[4])

                # Tagging
                keywords = f"{collection} {pattern} {description} {material} {name}"
                colors = color

                # Image
                thumbnail = row[25].replace("dl=0", "dl=1")
                roomsets = [row[id].replace("dl=0", "dl=1")
                            for id in range(26, 32) if row[id]]

                # Status
                statusP = mpn not in discontinuedMPNs
                statusS = False

                # Shipping
                shippingHeight = common.toFloat(row[20])
                shippingWidth = common.toFloat(row[21])
                shippingDepth = common.toFloat(row[22])
                shippingWeight = common.toFloat(row[19])
                if shippingWidth > 95 or shippingHeight > 95 or shippingDepth > 95 or shippingWeight > 40:
                    whiteGlove = True
                else:
                    whiteGlove = False

                leadtime = common.toText(row[17])
                if "days" in leadtime.lower() or "hours" in leadtime.lower() or "1 week" == leadtime.lower():
                    quickShip = True
                else:
                    quickShip = False

                # Stock
                if mpn in stocks:
                    stockP, stockNote = stocks[mpn]
                else:
                    stockP = 0
                    stockNote = ""

                # Fine-tuning
                TYPE_DICT = {
                    "tray": "Tray",
                    "ottoman": "Ottoman",
                    "stool": "Stool",
                    "bench": "Bench",
                    "pouf": "Pouf",
                    "dresser": "Dresser",
                    "nighstand": "Dresser",
                    "counter": "Counter Stool",
                    "counter stool": "Counter Stool",
                    "counterstool": "Counter Stool",
                    "bar stool": "Bar Stool",
                    "barstool": "Bar Stool",
                    "buffet": "Buffet",
                    "sideboard": "Sideboard",
                    "headboard": "Headboard",
                    "cube": "Accent Chair",
                    "chair": "Chair",
                    "seat": "Chair",
                    "dining chair": "Dining Chair",
                    "table": "Accent Table",
                    "dining table": "Dining Table",
                    "coffee table": "Coffee Table",
                    "sofa": "Sofa",
                    "console": "Console",
                }
                for key in TYPE_DICT.keys():
                    if key in pattern.lower():
                        type = TYPE_DICT[key]

                if "," in name:
                    pattern = name.split(",")[0].strip()
                    color = name.split(",")[1].strip()

                name = f"{pattern} {color.replace(',', '')}"

                # Exceptions
                if cost == 0 or not pattern or not color or not type:
                    continue

            except Exception as e:
                debug.warn(BRAND, str(e))
                continue

            product = {
                'mpn': mpn,
                'sku': sku,
                'pattern': pattern,
                'color': color,
                'name': name,

                'brand': brand,
                'type': type,
                'manufacturer': manufacturer,
                'collection': collection,

                'description': description,
                'width': width,
                'length': length,
                'height': height,

                'material': material,
                'country': country,
                'weight': weight,

                'specs': specs,
                'features': features,

                'uom': uom,

                'cost': cost,
                'map': map,

                'keywords': keywords,
                'colors': colors,

                'thumbnail': thumbnail,
                'roomsets': roomsets,

                'statusP': statusP,
                'statusS': statusS,

                'whiteGlove': whiteGlove,
                'quickShip': quickShip,

                'stockP': stockP,
                'stockNote': stockNote,
            }
            products.append(product)

        return products

    def inventory(self):
        stocks = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/peninsulahome-inventory.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = common.toText(row[0])

            try:
                product = PeninsulaHome.objects.get(mpn=mpn)
            except PeninsulaHome.DoesNotExist:
                continue

            sku = product.sku
            stockP = common.toInt(row[1])
            stockNote = common.toText(row[2])

            stock = {
                'sku': sku,
                'quantity': stockP,
                'note': stockNote
            }
            stocks.append(stock)

        return stocks
