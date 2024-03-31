from django.core.management.base import BaseCommand
from feed.models import HubbardtonForge

import os
import openpyxl

from utils import database, debug, common

BRAND = "Hubbardton Forge"
FILEDIR = f"{os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}/files"


class Command(BaseCommand):
    help = f"Build {BRAND} Database"

    def add_arguments(self, parser):
        parser.add_argument('functions', nargs='+', type=str)

    def handle(self, *args, **options):
        if "feed" in options['functions']:
            processor = Processor()
            feeds = processor.fetchFeed()
            processor.DatabaseManager.writeFeed(feeds=feeds)

        if "validate" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.validateFeed()

        if "status" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.statusSync(fullSync=False)

        if "content" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.contentSync()

        if "price" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.priceSync()

        if "tag" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.tagSync()

        if "add" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.addProducts()

        if "update" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.updateProducts(
                feeds=HubbardtonForge.objects.all(), private=False)

        if "image" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.downloadImages()

        if "inventory" in options['functions']:
            processor = Processor()
            stocks = processor.inventory()
            processor.DatabaseManager.updateInventory(
                stocks=stocks, type=3, reset=True)


class Processor:
    def __init__(self):
        self.DatabaseManager = database.DatabaseManager(
            brand=BRAND, Feed=HubbardtonForge)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def fetchFeed(self):
        # Price & Discontinued
        prices = {}

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/hubbardtonforge-price.xlsx", data_only=True)
        sh = wb.worksheets[2]
        for row in sh.iter_rows(min_row=3, values_only=True):
            mpn = common.toText(row[3])

            cost = common.toFloat(row[5])
            map = common.toFloat(row[6])

            prices[mpn] = {
                'cost': cost,
                'map': map,
            }

        # Get Product Feed
        products = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/hubbardtonforge-master.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            try:
                # Primary Keys
                mpn = common.toText(row[1])
                sku = f"HF {mpn}"

                pattern = common.toInt(row[3])

                colorOptions = [9, 10, 11, 12, 13]
                color = ' '.join(common.toText(
                    row[i]) for i in colorOptions if common.toText(row[i]))

                name = f"{color} {common.toText(row[2])}"

                # Categorization
                brand = BRAND
                type = common.toText(row[4])
                manufacturer = BRAND
                collection = common.toText(row[7])

                # Main Information
                description = f"{common.toText(row[72])} {common.toText(row[73])}"

                width = common.toFloat(row[21])
                length = common.toFloat(row[22])
                height = common.toFloat(row[20])

                # Additional Information
                finish = ', '.join(common.toText(
                    row[i]) for i in colorOptions if common.toText(row[i]))

                weight = common.toFloat(row[26])

                features = []
                for id in range(74, 79):
                    if row[id]:
                        features.append(common.toText(row[id]))

                # Measurement
                uom = "Item"

                # Pricing
                cost = common.toFloat(row[17])
                map = common.toFloat(row[18])
                msrp = common.toFloat(row[19])

                if mpn in prices:
                    cost = prices[mpn]['cost']
                    map = prices[mpn]['map']

                # Tagging
                keywords = f"{collection} {pattern} {description} {row[79]}, {row[80]}, {row[89]}, {type}, {name}, {finish}"
                colors = color

                # Image
                thumbnail = row[90]

                # Status
                statusP = True
                statusS = False

                # Fine-tuning
                TYPE_DICT = {
                    "Chandeliers": "Chandelier",
                    "Kitchen Pendants": "Pendant",
                    "Pendants": "Pendant",
                    "Semi-Flush": "Semi-Flush Mount",
                    "Large Scale Fixtures": "Accessory",
                    "Sconces - Direct": "Wall Sconce",
                    "Floor Lamps": "Floor Lamp",
                    "Torchieres": "Torchier",
                    "Table Lamps": "Table Lamp",
                    "Sconces - Pinup": "Wall Sconce",
                    "Outdoor": "Accessory",
                    "Home Accessories": "Accessory"
                }
                type = TYPE_DICT.get(type, type)

                # Exceptions
                if cost == 0 or not pattern or not color or not type:
                    continue

            except Exception as e:
                debug.warn(BRAND, str(e))
                continue

            product = {
                'mpn': mpn,
                'sku': sku,
                'pattern': pattern,
                'color': color,
                'name': name,

                'brand': brand,
                'type': type,
                'manufacturer': manufacturer,
                'collection': collection,

                'description': description,
                'width': width,
                'length': length,
                'height': height,

                'finish': finish,
                'weight': weight,
                'features': features,

                'uom': uom,

                'cost': cost,
                'map': map,
                'msrp': msrp,

                'keywords': keywords,
                'colors': colors,

                'thumbnail': thumbnail,

                'statusP': statusP,
                'statusS': statusS,
            }
            products.append(product)

        return products

    def inventory(self):
        stocks = []

        feeds = HubbardtonForge.objects.all()
        for feed in feeds:
            stock = {
                'sku': feed.sku,
                'quantity': 5,
                'note': "Made To Order: This product will ship in 3-4 weeks."
            }
            stocks.append(stock)

        return stocks
