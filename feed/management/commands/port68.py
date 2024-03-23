from django.core.management.base import BaseCommand
from feed.models import Port68

import os
import openpyxl

from utils import database, debug, common

BRAND = "Port 68"
FILEDIR = f"{os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}/files"


class Command(BaseCommand):
    help = f"Build {BRAND} Database"

    def add_arguments(self, parser):
        parser.add_argument('functions', nargs='+', type=str)

    def handle(self, *args, **options):
        if "feed" in options['functions']:
            processor = Processor()
            feeds = processor.fetchFeed()
            processor.DatabaseManager.writeFeed(feeds=feeds)

        if "validate" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.validateFeed()

        if "status" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.statusSync(fullSync=True)

        if "content" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.contentSync()

        if "price" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.priceSync()

        if "tag" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.tagSync()

        if "add" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.addProducts()

        if "update" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.updateProducts(
                feeds=Port68.objects.all(), private=False)

        if "image" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.downloadImages()

        if "inventory" in options['functions']:
            common.downloadFileFromSFTP(
                brand=BRAND,
                src="/port68",
                dst=f"{FILEDIR}/port68-inventory.xlsx",
                fileSrc=False,
                delete=True
            )

            processor = Processor()
            stocks = processor.inventory()
            processor.DatabaseManager.updateInventory(
                stocks=stocks, type=1, reset=True)


class Processor:
    def __init__(self):
        self.DatabaseManager = database.DatabaseManager(
            brand=BRAND, Feed=Port68)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def fetchFeed(self):
        # Get Product Feed
        products = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/port68-master.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            try:
                # Primary Keys
                mpn = common.toText(row[2])
                sku = f"P68 {mpn}"

                pattern = common.toText(row[3])
                color = common.toText(row[4])

                # Categorization
                brand = BRAND
                manufacturer = BRAND

                type = common.toText(row[5])

                collection = common.toText(row[1])

                # Main Information
                description = common.toText(row[19])

                # Additional Information
                material = common.toText(row[12])
                care = common.toText(row[25])
                country = common.toText(row[35])
                weight = common.toFloat(row[14])
                upc = common.toText(row[13])

                dimension = common.toText(row[18]).replace("Size:", "").strip()
                specs = [
                    ("Dimension", dimension)
                ]

                features = [common.toText(row[index]) for index in (
                    20, 21, 22, 23, 27) if row[index]]

                # Measurement
                uom = "Item"

                # Pricing
                cost = common.toFloat(row[7])
                map = common.toFloat(row[8])
                msrp = common.toFloat(row[9])

                # Tagging
                keywords = f"{row[11]} {type} {description} {row[20]}"
                colors = color

                # Image
                thumbnail = row[51]
                roomsets = [row[id] for id in range(52, 65) if row[id] != ""]

                # Status
                statusP = True
                statusS = False

                # Fine-tuning
                type = common.pluralToSingular(type)
                TYPE_DICT = {
                    "Jar": "Ginger Jar",
                    "Urn": "Ginger Jar",
                    "Accent Table/Tray": "Tray",
                    "Buffet Lamp": "Accent Lamp",
                    "Lamp": "Accent Lamp",
                }
                type = TYPE_DICT.get(type, type)

                MANUFACTURER_DICT = {
                    "Scalamandre": "Scalamandre Maison",
                    "Madcap Cottage": "Madcap Cottage Décor",
                    "Williamsburg": "Williamsburg"
                }
                manufacturer = MANUFACTURER_DICT.get(collection)

                name = f"{pattern} {color} {type}"

                # Exceptions
                if cost == 0 or not pattern or not color or not type:
                    continue

            except Exception as e:
                debug.warn(BRAND, str(e))
                continue

            product = {
                'mpn': mpn,
                'sku': sku,
                'pattern': pattern,
                'color': color,
                'name': name,

                'brand': brand,
                'type': type,
                'manufacturer': manufacturer,
                'collection': collection,

                'description': description,

                'material': material,
                'care': care,
                'country': country,
                'weight': weight,
                'upc': upc,

                'specs': specs,
                'features': features,

                'uom': uom,

                'cost': cost,
                'map': map,
                'msrp': msrp,

                'keywords': keywords,
                'colors': colors,

                'thumbnail': thumbnail,
                'roomsets': roomsets,

                'statusP': statusP,
                'statusS': statusS,
            }
            products.append(product)

        return products

    def inventory(self):
        stocks = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/port68-inventory.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = common.toText(row[0])

            try:
                product = Port68.objects.get(mpn=mpn)
            except Port68.DoesNotExist:
                continue

            sku = product.sku
            stockP = common.toInt(row[2])
            stockNote = common.toText(row[3])

            stock = {
                'sku': sku,
                'quantity': stockP,
                'note': stockNote
            }
            stocks.append(stock)

        return stocks
