from django.core.management.base import BaseCommand
from feed.models import DanaGibson

import os
import openpyxl
import re

from utils import database, debug, common

BRAND = "Dana Gibson"
FILEDIR = f"{os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}/files"


class Command(BaseCommand):
    help = f"Build {BRAND} Database"

    def add_arguments(self, parser):
        parser.add_argument('functions', nargs='+', type=str)

    def handle(self, *args, **options):
        if "feed" in options['functions']:
            processor = Processor()
            feeds = processor.fetchFeed()
            processor.DatabaseManager.writeFeed(feeds=feeds)

        if "validate" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.validateFeed()

        if "status" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.statusSync(fullSync=False)

        if "content" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.contentSync()

        if "price" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.priceSync()

        if "tag" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.tagSync()

        if "add" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.addProducts()

        if "update" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.updateProducts(
                feeds=DanaGibson.objects.all(), private=False)

        if "image" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.downloadImages()

        if "inventory" in options['functions']:
            processor = Processor()
            stocks = processor.inventory()
            processor.DatabaseManager.updateInventory(
                stocks=stocks, type=1, reset=True)


class Processor:
    def __init__(self):
        self.DatabaseManager = database.DatabaseManager(
            brand=BRAND, Feed=DanaGibson)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def fetchFeed(self):
        # Get Product Feed
        products = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/danagibson-master.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            try:
                # Primary Keys
                mpn = common.toText(row[1])
                sku = f"DG {mpn}"

                pattern = common.toText(row[4]).title()
                color = common.toText(row[3]).title()

                # Categorization
                brand = BRAND
                manufacturer = BRAND
                type = common.toText(row[2]).title()
                collection = common.toText(row[2])

                # Main Information
                description = common.toText(row[17])
                width = common.toFloat(row[12])
                length = common.toFloat(row[10])
                height = common.toFloat(row[14])

                # Additional Information
                material = common.toText(row[19])
                finish = common.toText(row[20])
                country = common.toText(row[32])
                weight = common.toFloat(row[9])

                specs = [(common.toText([j]).title(), common.toText(row[j]))
                         for j in [21, 22, 26, 27, 28, 29] if common.toText(row[j])]

                # Pricing
                cost = common.toFloat(row[5])
                map = common.toFloat(row[6])

                # Measurement
                uom = "Item"

                # Tagging
                keywords = f"{row[18]}, {pattern}, {type}"
                colors = color

                # Image
                thumbnail = row[46]
                roomsets = [row[id] for id in range(52, 54) if row[id]]

                # Status
                if common.toText(row[1]).lower() == "oos":
                    statusP = False
                else:
                    statusP = True

                statusS = False

                # Fine-tuning
                pattern = pattern.replace(color, "").replace(
                    type, "").replace("Lamp", "").replace("  ", " ").strip()

                TYPE_DICT = {
                    "Sconce": "Wall Sconce",
                    "Tablelamp": "Table Lamp",
                    "Lumbar": "Pillow",
                    "Boudoir": "Pillow",
                    "Wastebsaket": "Wastebasket",
                    "Waste": "Wastebasket",
                    "Wastrebasket": "Wastebasket",
                    "Table Lamps": "Table Lamp",
                    "Large Cachepot": "Cachepot",
                    "Small Cachepot": "Cachepot",
                    "Small Ginger Jar": "Ginger Jar",
                    "Large Lachepot": "Accent",
                    "Long Cachepot": "Cachepot",
                    "Candlesticks": "Candlestick",
                }
                type = TYPE_DICT.get(type, type)

                name = common.toText(row[4]).title()
                if "Lumbar" in name:
                    name = f"{name} Pillow"

                # Exceptions
                if cost == 0 or not pattern or not color or not type:
                    continue

            except Exception as e:
                debug.warn(BRAND, str(e))
                continue

            product = {
                'mpn': mpn,
                'sku': sku,
                'pattern': pattern,
                'color': color,
                'name': name,

                'brand': brand,
                'type': type,
                'manufacturer': manufacturer,
                'collection': collection,

                'description': description,
                'width': width,
                'length': length,
                'height': height,

                'material': material,
                'finish': finish,
                'country': country,
                'weight': weight,

                'specs': specs,

                'uom': uom,

                'cost': cost,
                'map': map,

                'keywords': keywords,
                'colors': colors,

                'thumbnail': thumbnail,
                'roomsets': roomsets,

                'statusP': statusP,
                'statusS': statusS,
            }
            products.append(product)

        return products

    def inventory(self):
        stocks = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/danagibson-inventory.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = row[1]

            try:
                product = DanaGibson.objects.get(mpn=mpn)
            except DanaGibson.DoesNotExist:
                continue

            sku = product.sku
            stockP = common.toInt(row[3])

            stock = {
                'sku': sku,
                'quantity': stockP,
                'note': "6 days"
            }
            stocks.append(stock)

        return stocks
