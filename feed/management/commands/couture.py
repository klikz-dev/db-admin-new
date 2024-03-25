from django.core.management.base import BaseCommand
from feed.models import Couture

import os
import environ
import openpyxl
import re

from utils import database, debug, common

env = environ.Env()

BRAND = "Couture"
FILEDIR = f"{os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}/files"


class Command(BaseCommand):
    help = f"Build {BRAND} Database"

    def add_arguments(self, parser):
        parser.add_argument('functions', nargs='+', type=str)

    def handle(self, *args, **options):
        if "feed" in options['functions']:
            processor = Processor()
            feeds = processor.fetchFeed()
            processor.DatabaseManager.writeFeed(feeds=feeds)

        if "validate" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.validateFeed()

        if "status" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.statusSync(fullSync=True)

        if "content" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.contentSync()

        if "price" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.priceSync()

        if "tag" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.tagSync()

        if "add" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.addProducts()

        if "update" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.updateProducts(
                feeds=Couture.objects.all())

        if "image" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.downloadImages()

        if "inventory" in options['functions']:
            common.downloadFileFromSFTP(
                brand=BRAND,
                src="/couture",
                dst=f"{FILEDIR}/couture-inventory.xlsm",
                fileSrc=False,
                delete=False
            )

            processor = Processor()
            stocks = processor.inventory()
            processor.DatabaseManager.updateInventory(
                stocks=stocks, type=1, reset=True)


class Processor:
    def __init__(self):
        self.DatabaseManager = database.DatabaseManager(
            brand=BRAND, Feed=Couture)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def fetchFeed(self):
        # Get Product Feed
        products = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/couture-master.xlsx", data_only=True)
        sh = wb.worksheets[0]

        header = []
        for row in sh.iter_rows(min_row=2, max_row=2, values_only=True):
            header = row

        for row in sh.iter_rows(min_row=3, values_only=True):
            try:
                # Primary Keys
                mpn = common.toText(row[0])
                sku = f"CL {mpn}"

                pattern = common.toText(row[2].split("-")[0])
                color = common.toText(row[7])

                # Categorization
                brand = BRAND
                manufacturer = BRAND
                type = common.toText(row[3]).title()
                collection = type

                # Main Information
                description = common.toText(row[4])
                width = common.toFloat(row[13])
                length = common.toFloat(row[16])
                height = common.toFloat(row[15])

                # Additional Information
                material = common.toText(row[6])
                care = common.toText(row[5])
                country = common.toText(row[1])
                weight = common.toFloat(row[12])
                upc = row[38]

                features = [common.toText(row[8])]

                specs = [(common.toText(header[j]).title(), common.toText(row[j]))
                         for j in [17, 19, 20, 24, 25, 26] if common.toText(row[j])]

                # Measurement
                uom = "Item"

                # Pricing
                cost = common.toFloat(row[9])
                map = common.toFloat(row[10])

                # Tagging
                keywords = description
                colors = color

                # Status
                statusP = True
                statusS = False

                # Fine-tuning
                TYPE_DICT = {
                    "Decorative Accessories": "Decorative Accent",
                    "Pendants": "Pendant",
                    "Floor Lamps": "Floor Lamp",
                }
                type = TYPE_DICT.get(type, type)

                pattern = re.sub(r"Lamp|Table|\s{2,}", "", pattern).strip()

                name = f"{pattern} {color} {type}"

                # Exceptions
                if cost == 0 or not pattern or not color or not type:
                    continue

            except Exception as e:
                debug.warn(BRAND, str(e))
                continue

            product = {
                'mpn': mpn,
                'sku': sku,
                'pattern': pattern,
                'color': color,
                'name': name,

                'brand': brand,
                'type': type,
                'manufacturer': manufacturer,
                'collection': collection,

                'description': description,
                'width': width,
                'length': length,
                'height': height,

                'material': material,
                'care': care,
                'country': country,
                'weight': weight,
                'upc': upc,

                'features': features,
                'specs': specs,

                'uom': uom,

                'cost': cost,
                'map': map,

                'keywords': keywords,
                'colors': colors,

                'statusP': statusP,
                'statusS': statusS,
            }
            products.append(product)

        return products

    def inventory(self):
        stocks = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/couture-inventory.xlsm", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = common.toText(row[0])

            try:
                product = Couture.objects.get(mpn=mpn)
            except Couture.DoesNotExist:
                continue

            sku = product.sku

            stockP = common.toInt(row[1])
            stockNote = row[2]

            stock = {
                'sku': sku,
                'quantity': stockP,
                'note': stockNote
            }
            stocks.append(stock)

        return stocks
