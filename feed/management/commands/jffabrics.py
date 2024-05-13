from django.core.management.base import BaseCommand
from feed.models import JFFabrics

import os
import environ
import openpyxl

from utils import database, debug, common

env = environ.Env()

BRAND = "JF Fabrics"
FILEDIR = f"{os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}/files"


class Command(BaseCommand):
    help = f"Build {BRAND} Database"

    def add_arguments(self, parser):
        parser.add_argument('functions', nargs='+', type=str)

    def handle(self, *args, **options):
        if "feed" in options['functions']:
            processor = Processor()
            feeds = processor.fetchFeed()
            processor.DatabaseManager.writeFeed(feeds=feeds)

        if "validate" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.validateFeed()

        if "status" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.statusSync(fullSync=False)

        if "content" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.contentSync()

        if "price" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.priceSync()

        if "tag" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.tagSync()

        if "add" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.addProducts()

        if "image" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.downloadImages()

        if "inventory" in options['functions']:
            common.downloadFileFromSFTP(
                brand=BRAND,
                src="/Decorating Best Inventory.xlsx",
                dst=f"{FILEDIR}/jffabrics-inventory.xlsx",
                fileSrc=True,
                delete=False
            )

            processor = Processor()
            stocks = processor.inventory()
            processor.DatabaseManager.updateInventory(
                stocks=stocks, type=1, reset=True)


class Processor:
    def __init__(self):
        self.DatabaseManager = database.DatabaseManager(
            brand=BRAND, Feed=JFFabrics)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def fetchFeed(self):

        # Disco Books
        discoBooks = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/jffabrics-disco-books.xlsx", data_only=True)
        sh = wb.worksheets[0]
        for row in sh.iter_rows(min_row=3, values_only=True):
            book = common.toText(row[0])
            discoBooks.append(book)

        # Disco Skus
        discoMPNs = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/jffabrics-disco-skus.xlsx", data_only=True)
        sh = wb.worksheets[0]
        for row in sh.iter_rows(min_row=3, values_only=True):
            book = common.toText(row[3])
            pattern = common.toText(row[0])
            color = common.toInt(row[1])

            mpn = f"{pattern}_{color}{book}"
            discoMPNs.append(mpn)

        # Get Product Feed
        products = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/jffabrics-master.xlsx", data_only=True)
        sh = wb.worksheets[0]
        for row in sh.iter_rows(min_row=2, values_only=True):
            try:
                # Primary Keys
                book = common.toText(row[1])
                pattern = common.toText(row[2])
                color = common.toInt(row[3])

                mpn = f"{pattern}_{color}{book}"
                sku = f"JF {common.toInt(row[4])}"

                # Categorization
                brand = BRAND
                manufacturer = BRAND
                type = common.toText(row[8])
                collection = book

                # Main Information
                description = common.toText(row[59])
                width = common.toFloat(row[24])
                repeatH = common.toFloat(row[30])
                repeatV = common.toFloat(row[31])

                # Additional Information
                content = common.toText(row[10])
                yardsPR = common.toFloat(row[25])
                usage = common.toText(row[13])
                country = common.toText(row[23])
                weight = common.toFloat(row[34])

                # Measurement
                uom = common.toText(row[73])

                # Pricing
                cost = common.toFloat(row[76])
                map = common.toFloat(row[74])

                # Tagging
                keywords = f"{collection} {pattern} {description} {row[14]} {row[22]}"
                colors = common.toText(row[6])

                # Image
                thumbnail = row[78]

                # Status
                statusP = True
                statusS = True

                # Fine-tuning
                TYPE_DICT = {
                    "Wallcovering": "Wallpaper"
                }
                type = TYPE_DICT.get(type, type)

                if type == "Fabric" and "Upholstery" in usage:
                    type = "Upholstery Fabric"
                elif type == "Fabric" and "Drapery" in usage:
                    type = "Drapery Fabric"

                UOM_DICT = {
                    "DR": "Roll",
                    "EA": "Yard",
                    "YD": "Yard",
                }
                uom = UOM_DICT.get(uom, uom)

                if colors:
                    color = f"{'/'.join([c.strip() for c in colors.split(',')])} ({color})"

                if "mural" in description.lower():
                    type = "Mural"

                name = f"{pattern} {color} {type}"

                # Exceptions
                if book in discoBooks or mpn in discoMPNs:
                    continue

                if cost == 0 or not pattern or not color or not type:
                    continue

            except Exception as e:
                debug.warn(BRAND, str(e))
                continue

            product = {
                'mpn': mpn,
                'sku': sku,
                'pattern': pattern,
                'color': color,
                'name': name,

                'brand': brand,
                'type': type,
                'manufacturer': manufacturer,
                'collection': collection,

                'description': description,
                'width': width,
                'repeatV': repeatV,
                'repeatH': repeatH,

                'yardsPR': yardsPR,
                'content': content,
                'usage': usage,
                'country': country,
                'weight': weight,

                'uom': uom,

                'cost': cost,
                'map': map,

                'keywords': keywords,
                'colors': colors,

                'thumbnail': thumbnail,

                'statusP': statusP,
                'statusS': statusS,
            }
            products.append(product)

        return products

    def inventory(self):
        stocks = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/jffabrics-inventory.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            sku = f"JF {common.toInt(row[0])}"

            try:
                product = JFFabrics.objects.get(sku=sku)
            except JFFabrics.DoesNotExist:
                continue

            sku = product.sku
            stockP = common.toInt(row[3])

            stock = {
                'sku': sku,
                'quantity': stockP,
                'note': ""
            }
            stocks.append(stock)

        return stocks
