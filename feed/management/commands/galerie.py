from django.core.management.base import BaseCommand
from feed.models import Galerie

import os
import environ
import openpyxl

from utils import database, debug, common

env = environ.Env()

BRAND = "Galerie"
FILEDIR = f"{os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}/files"


class Command(BaseCommand):
    help = f"Build {BRAND} Database"

    def add_arguments(self, parser):
        parser.add_argument('functions', nargs='+', type=str)

    def handle(self, *args, **options):
        if "feed" in options['functions']:
            processor = Processor()
            feeds = processor.fetchFeed()
            processor.DatabaseManager.writeFeed(feeds=feeds)

        if "validate" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.validateFeed()

        if "status" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.statusSync(fullSync=False)

        if "content" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.contentSync()

        if "price" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.priceSync()

        if "tag" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.tagSync()

        if "add" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.addProducts()

        if "update" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.updateProducts(
                feeds=Galerie.objects.all())

        if "image" in options['functions']:
            processor = Processor()
            processor.DatabaseManager.downloadImages()

        if "inventory" in options['functions']:
            common.downloadFileFromSFTP(
                brand=BRAND,
                src="/galerie/GalerieStock.xlsx",
                dst=f"{FILEDIR}/galerie-inventory.xlsx",
                fileSrc=True,
                delete=False
            )

            processor = Processor()
            stocks = processor.inventory()
            processor.DatabaseManager.updateInventory(
                stocks=stocks, type=1, reset=True)


class Processor:
    def __init__(self):
        self.DatabaseManager = database.DatabaseManager(
            brand=BRAND, Feed=Galerie)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def fetchFeed(self):
        # Get Product Feed
        products = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/galerie-master.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=3, values_only=True):
            try:
                # Primary Keys
                mpn = common.toText(row[2])
                sku = f"G {mpn}"

                pattern = common.toText(row[3])
                color = common.toText(row[4])

                # Categorization
                brand = BRAND
                manufacturer = BRAND
                type = "Wallpaper"
                collection = common.toText(row[1])

                # Main Information
                description = common.toText(row[8])
                width = common.toFloat(row[16])
                length = common.toFloat(row[17]) * 12
                repeatV = common.toFloat(row[20])
                repeatH = common.toFloat(row[21])

                # Additional Information
                yardsPR = common.toFloat(row[17] / 3)
                match = common.toText(row[22])
                material = common.toText(row[24])
                weight = common.toFloat(row[19])
                country = common.toText(row[29])

                coverage = common.toText(row[18])
                paste = common.toText(row[23])
                washability = common.toText(row[25])
                removability = common.toText(row[26])
                specs = [
                    ("Coverage", coverage),
                    ("Paste", paste),
                    ("Washability", washability),
                    ("Removability", removability),
                ]

                # Pricing
                cost = common.toFloat(row[9])
                map = common.toFloat(row[10])

                # Measurement
                uom = common.toText(row[12])

                # Tagging
                keywords = f"{collection} {pattern} {description} {match} {paste} {material} {washability} {removability} {common.toText(row[27])}"
                colors = color

                # Image
                thumbnail = row[30]
                roomsets = [row[id] for id in range(31, 35) if row[id] != ""]

                # Status
                statusP = True
                statusS = True

                # Fine-tuning
                UOM_DICT = {
                    "Mural": "Roll",
                }
                uom = UOM_DICT.get(uom, uom)

                name = f"{pattern} {color} {type}"

                # Exceptions
                if cost == 0 or not pattern or not color or not type:
                    continue

            except Exception as e:
                debug.warn(BRAND, str(e))
                continue

            product = {
                'mpn': mpn,
                'sku': sku,
                'pattern': pattern,
                'color': color,
                'name': name,

                'brand': brand,
                'type': type,
                'manufacturer': manufacturer,
                'collection': collection,

                'description': description,
                'width': width,
                'length': length,
                'repeatV': repeatV,
                'repeatH': repeatH,

                'yardsPR': yardsPR,
                'match': match,
                'material': material,
                'weight': weight,
                'country': country,

                'specs': specs,

                'uom': uom,

                'cost': cost,
                'map': map,

                'keywords': keywords,
                'colors': colors,

                'thumbnail': thumbnail,
                'roomsets': roomsets,

                'statusP': statusP,
                'statusS': statusS,
            }
            products.append(product)

        return products

    def inventory(self):
        stocks = []

        wb = openpyxl.load_workbook(
            f"{FILEDIR}/galerie-inventory.xlsx", data_only=True)
        sh = wb.worksheets[0]

        for row in sh.iter_rows(min_row=2, values_only=True):
            mpn = common.toText(row[0]).replace("âˆ’", "-").replace(".0", "")

            try:
                product = Galerie.objects.get(mpn=mpn)
            except Galerie.DoesNotExist:
                continue

            sku = product.sku
            stockP = common.toInt(row[3])
            stockNote = common.toText(row[2])

            stock = {
                'sku': sku,
                'quantity': stockP,
                'note': stockNote
            }
            stocks.append(stock)

        return stocks
